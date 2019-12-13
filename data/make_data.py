import openpyxl
import tqdm
import csv


book = openpyxl.load_workbook('simple_data.xlsx')
sheet = book['平易化コーパス']
with open('simple_data.css', 'w') as f:
    for i in tqdm.tqdm(range(1, 35000)):
        writer = csv.writer(f)
        writer.writerow([sheet.cell(row=i, column=2).value, sheet.cell(row=i, column=3).value, sheet.cell(row=i, column=5).value])
