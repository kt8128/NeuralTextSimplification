import openpyxl
import tqdm
import csv


book = openpyxl.load_workbook('simple_data.xlsx')
sheet = book['評価用']
with open('difficult-easy-correct.csv', 'w') as f:
    for i in tqdm.tqdm(range(1, 101)):
        writer = csv.writer(f)
        writer.writerow([sheet.cell(row=i, column=2).value, sheet.cell(row=i, column=3).value])
